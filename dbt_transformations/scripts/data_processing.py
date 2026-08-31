"""
Pré-processamento: converte Excel brutos (data/raw/) em Parquet (data/processed/).

Uso: python scripts/data_processing.py

Fluxo completo: Excel → Parquet → dbt (DuckDB) → Metabase
"""

import logging
import time
import unicodedata
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# Diretórios
RAIZ = Path(__file__).resolve().parent.parent.parent
RAW_DIR = RAIZ / "data" / "raw"
PROCESSED_DIR = RAIZ / "data" / "processed"

# Fontes Excel
EXCEL_DTB = RAW_DIR / "DTB_relatorio_br_municipios_2024.xls"
EXCEL_PIB = RAW_DIR / "PIB_municípios_2010-2023.xlsx"
EXCEL_PROJ = RAW_DIR / "projetos_IA_2020-2025.xlsx"
EXCEL_IDEB = RAW_DIR / "IDEB_anos_iniciais_2023.xlsx"

# Mapeamento IBGE: código numérico → sigla da UF (fixo, não muda)
UF_SIGLAS = {
    "11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA",
    "16": "AP", "17": "TO", "21": "MA", "22": "PI", "23": "CE",
    "24": "RN", "25": "PB", "26": "PE", "27": "AL", "28": "SE",
    "29": "BA", "31": "MG", "32": "ES", "33": "RJ", "35": "SP",
    "41": "PR", "42": "SC", "43": "RS", "50": "MS", "51": "MT",
    "52": "GO", "53": "DF",
}

# Tabela de-para: nomes sujos em projetos_ia → nome oficial do DTB
# Adicione aqui toda exceção nova que o assert abaixo acusar.
DEPARA_MUNICIPIO_SUJO = {
    ("PB", "CAMPINA GRANDE- MIXING CENTER"): "CAMPINA GRANDE",
    ("PB", "QUEIMADAS *"): "QUEIMADAS",
}

# UFs válidas no parquet de projetos_ia (filtra lixo de cabeçalho)
UF_VALIDAS_PROJETOS = {"MG", "PB", "PE", "RJ", "SP"}


def _dedup_colunas(nomes):
    """Gera nomes de coluna únicos com sufixo .N para duplicatas.

    Replica o comportamento do Pandas ao concatenar DataFrames com
    colunas homônimas: a primeira ocorrência fica sem sufixo, as
    demais recebem .1, .2, etc. Colunas vazias viram 'Unnamed: N'.
    """
    visto = {}
    resultado = []
    contador_unnamed = 0
    for n in nomes:
        texto = "" if n is None else str(n).strip()
        if texto == "" or texto == "None":
            resultado.append(f"Unnamed: {contador_unnamed}")
            contador_unnamed += 1
        else:
            if texto in visto:
                visto[texto] += 1
                resultado.append(f"{texto}.{visto[texto]}")
            else:
                visto[texto] = 0
                resultado.append(texto)
    return resultado


def _ler_excel_sem_cabecalho(caminho, linha_cabecalho):
    """Lê Excel preservando os nomes originais das colunas.

    O engine calamine (default do Polars) dropa linhas vazias
    automaticamente, deslocando os índices. Esta função lê sem
    cabeçalho, seleciona a linha correta e renomeia manualmente.
    """
    df_raw = pl.read_excel(caminho, has_header=False, infer_schema_length=0)
    assert df_raw.height > linha_cabecalho, f"Arquivo muito curto: {caminho.name}"

    cabecalho = [str(v) for v in df_raw.row(linha_cabecalho)]
    cabecalho = _dedup_colunas(cabecalho)

    df = df_raw.slice(linha_cabecalho + 1).rename(
        dict(zip(df_raw.columns, cabecalho))
    )
    assert df.height > 0, f"Arquivo vazio: {caminho.name}"
    return df.cast(pl.Utf8)


def _norm_str(s: str) -> str:
    """Normaliza string para comparação: uppercase, sem acentos, sem espaços nas pontas."""
    s = s.upper().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


def enriquecer_com_id_municipio(
    df_projetos: pl.DataFrame, df_dtb: pl.DataFrame
) -> pl.DataFrame:
    """Adiciona id_municipio (código IBGE 7 dígitos) via nome_municipio + sigla_uf.

    Estratégia: normalização → match exato → tabela de-para para exceções.
    Órfãos recebem id_municipio = "-1" com warning no log.
    """
    # Chave de lookup a partir do DTB
    dtb_chave = df_dtb.select(
        pl.col("sigla_uf"),
        pl.col("Nome_Município").map_elements(_norm_str, return_dtype=pl.Utf8).alias(
            "_chave_nome"
        ),
        pl.col("Código Município Completo").alias("id_municipio"),
    ).unique()

    # Tabela de-para como DataFrame auxiliar
    depara = pl.DataFrame(
        [
            (uf, _norm_str(nome_sujo), _norm_str(nome_oficial))
            for (uf, nome_sujo), nome_oficial in DEPARA_MUNICIPIO_SUJO.items()
        ],
        schema=["ESTADO", "_chave_nome_bruta", "_chave_nome_corrigida"],
        orient="row",
    )

    df = (
        df_projetos
        .with_columns(
            pl.col("CIDADES")
            .map_elements(_norm_str, return_dtype=pl.Utf8)
            .alias("_chave_nome_bruta")
        )
        .join(
            depara, on=["ESTADO", "_chave_nome_bruta"], how="left"
        )
        .with_columns(
            pl.coalesce(["_chave_nome_corrigida", "_chave_nome_bruta"]).alias(
                "_chave_nome"
            )
        )
        .join(dtb_chave, left_on=["ESTADO", "_chave_nome"], right_on=["sigla_uf", "_chave_nome"], how="left")
    )

    orfaos = df.filter(pl.col("id_municipio").is_null())
    if orfaos.height > 0:
        logging.warning(
            f"  ⚠ {orfaos.height} linha(s) sem id_municipio — receberão -1. "
            f"Adicione ao DEPARA_MUNICIPIO_SUJO:"
        )
        for row in orfaos.select("ESTADO", "CIDADES").unique().iter_rows():
            logging.warning(f"    {row}")

    df = df.with_columns(pl.col("id_municipio").fill_null("-1"))

    return df.drop("_chave_nome", "_chave_nome_bruta", "_chave_nome_corrigida")


def salvar(df, nome):
    """Salva DataFrame como Parquet em data/processed/."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    df.write_parquet(PROCESSED_DIR / f"{nome}.parquet")
    logging.info(f"  ✓ {nome}: {df.height:,} linhas × {df.width} colunas")


def main():
    inicio = time.perf_counter()
    logging.info("Convertendo Excel → Parquet...\n")

    # 1. DTB — Diretório Territorial Brasileiro
    #    Calamine dropa 2 linhas vazias → cabeçalho na linha 4 (0-indexed)
    #    Injeta sigla_uf a partir do código numérico (o DTB do IBGE não tem essa coluna)
    df_dtb = _ler_excel_sem_cabecalho(EXCEL_DTB, linha_cabecalho=4)
    df_dtb = df_dtb.with_columns(
        pl.col("UF").replace_strict(UF_SIGLAS, default=None).alias("sigla_uf")
    )
    salvar(df_dtb, "dtb_municipios")

    # 2. PIB — Produto Interno Bruto municipal (2010-2021)
    #    Sem metadado; cabeçalho na primeira linha
    df = pl.read_excel(EXCEL_PIB)
    df.columns = _dedup_colunas(df.columns)
    salvar(df.cast(pl.Utf8), "pib_municipios")

    # 3. Projetos IA — cada aba = um ano (2020..2025)
    #    Abas têm colunas diferentes; lê sem cabeçalho para preservar nomes
    #    originais e aplica dedup com sufixo .N (compatível com Pandas)
    wb = load_workbook(EXCEL_PROJ, read_only=True)
    abas_numericas = [a for a in wb.sheetnames if a.isdigit()]
    wb.close()

    dfs = []
    for aba in abas_numericas:
        df_raw = pl.read_excel(
            EXCEL_PROJ, sheet_name=aba, has_header=False, infer_schema_length=0
        )
        cabecalho = _dedup_colunas([str(v) for v in df_raw.row(0)])
        df = df_raw.slice(1).rename(dict(zip(df_raw.columns, cabecalho)))
        df = df.cast(pl.Utf8).with_columns(pl.lit(int(aba)).alias("ano"))
        dfs.append(df)

    df_projetos = pl.concat(dfs, how="diagonal")

    # Filtrar linhas com ESTADO inválido (lixo de cabeçalho de abas)
    df_projetos = df_projetos.filter(
        pl.col("ESTADO").is_in(list(UF_VALIDAS_PROJETOS))
    )

    # Enriquecer com id_municipio via DTB
    df_projetos = enriquecer_com_id_municipio(df_projetos, df_dtb)

    salvar(df_projetos, "projetos_ia")

    # 4. IDEB — Índice de Desenvolvimento da Educação Básica
    #    Cabeçalho real com nomes de coluna (SG_UF, CO_MUNICIPIO, etc.) na linha 7
    salvar(_ler_excel_sem_cabecalho(
        EXCEL_IDEB, linha_cabecalho=7), "ideb_municipios")

    # 5. TDI — Taxa de Distorção Idade-Série (1 arquivo por ano)
    #    Calamine dropa 2 linhas vazias → cabeçalho na linha 6 (0-indexed)
    arquivos = sorted(RAW_DIR.glob("TDI_MUNICIPIOS_*.xlsx"))
    assert arquivos, f"Nenhum arquivo TDI encontrado em {RAW_DIR}"

    dfs = []
    for arq in arquivos:
        df_raw = pl.read_excel(arq, has_header=False, infer_schema_length=0)
        cabecalho = _dedup_colunas([str(v) for v in df_raw.row(6)])
        df = df_raw.slice(7).rename(dict(zip(df_raw.columns, cabecalho)))
        dfs.append(df.cast(pl.Utf8))

    salvar(pl.concat(dfs), "taxa_distorcao")

    logging.info(f"\n✅ Concluído em {time.perf_counter() - inicio:.1f}s")


if __name__ == "__main__":
    main()
